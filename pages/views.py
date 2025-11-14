# pages/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.db.models import Q, Count
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Item, Rental, Notification, CustomUser
from .forms import (
    CustomUserCreationForm,
    CustomAuthenticationForm,
    ItemForm,
    ItemSearchForm,
    RentalRequestForm,
    UserProfileEditForm
)

def landing_view(request):
    """Beautiful landing page for all users - main entry point"""
    return render(request, 'landing.html')

def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('home')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('home')

from django.contrib.auth.decorators import login_required

from django.db.models import Q
from .forms import ItemSearchForm
from django.utils.timezone import datetime

@login_required(login_url='login')
def home_page_view(request):
    form = ItemSearchForm(request.GET)
    items = Item.objects.all().order_by('-date_posted')

    if form.is_valid():
        # Search by keyword
        if search_query := form.cleaned_data.get('search'):
            items = items.filter(
                Q(name__icontains=search_query) |
                Q(description__icontains=search_query)
            )

        # Filter by category
        if category := form.cleaned_data.get('category'):
            items = items.filter(category=category)

        # Filter by availability
        if form.cleaned_data.get('availability'):
            items = items.filter(is_available=True)

        # Sort items
        if sort_by := form.cleaned_data.get('sort_by'):
            if sort_by == 'price_asc':
                items = items.order_by('price')
            elif sort_by == 'price_desc':
                items = items.order_by('-price')
            elif sort_by == 'date_desc':
                items = items.order_by('-date_posted')
            elif sort_by == 'date_asc':
                items = items.order_by('date_posted')
    else:
        # Default sorting by newest first
        items = items.order_by('-date_posted')

    context = {
        'all_items': items,
        'form': form
    }
    return render(request, 'index.html', context)

@login_required(login_url='login')
def item_detail_view(request, pk):
    item = Item.objects.get(pk=pk) # Get one item by its primary key (pk)
    context = {
        'item': item,
    }
    return render(request, 'item_detail.html', context)



@login_required(login_url='login')
def item_list_view(request):
    form = ItemSearchForm(request.GET)
    items = Item.objects.all()

    if form.is_valid():
        # Search by keyword
        if search_query := form.cleaned_data.get('search'):
            items = items.filter(
                Q(name__icontains=search_query) |
                Q(description__icontains=search_query)
            )

        # Filter by category
        if category := form.cleaned_data.get('category'):
            items = items.filter(category=category)

        # Filter by availability
        if form.cleaned_data.get('availability'):
            items = items.filter(is_available=True)

        # Filter by date range
        if date_from := form.cleaned_data.get('date_from'):
            items = items.filter(date_posted__gte=datetime.combine(date_from, datetime.min.time()))
        if date_to := form.cleaned_data.get('date_to'):
            items = items.filter(date_posted__lte=datetime.combine(date_to, datetime.max.time()))

        # Sort items
        if sort_by := form.cleaned_data.get('sort_by'):
            if sort_by == 'price_asc':
                items = items.order_by('price')
            elif sort_by == 'price_desc':
                items = items.order_by('-price')
            elif sort_by == 'date_desc':
                items = items.order_by('-date_posted')
            elif sort_by == 'date_asc':
                items = items.order_by('date_posted')

    context = {
        'items': items,
        'form': form
    }
    return render(request, 'item_list.html', context)

@login_required(login_url='login')
def create_item_view(request):
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.owner = request.user
            item.save()
            return redirect('item_detail', pk=item.pk)
    else:
        form = ItemForm()

    context = {'form': form}
    return render(request, 'item_form.html', context)

@login_required(login_url='login')
def update_item_view(request, pk):
    item = get_object_or_404(Item, pk=pk)
    # Only the owner or a superuser can update
    if not (request.user == item.owner or request.user.is_superuser):
        raise PermissionDenied()
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            return redirect('item_detail', pk=item.pk)
    else:
        form = ItemForm(instance=item)
    
    context = {'form': form}
    return render(request, 'item_form.html', context)

@login_required(login_url='login')
def delete_item_view(request, pk):
    item = get_object_or_404(Item, pk=pk)
    # Only the owner or a superuser can delete
    if not (request.user == item.owner or request.user.is_superuser):
        raise PermissionDenied()
    if request.method == 'POST':
        item.delete()
        return redirect('home')
    
    context = {'item': item}
    return render(request, 'item_confirm_delete.html', context)

@login_required(login_url='login')
def rental_request_view(request, pk):
    item = get_object_or_404(Item, pk=pk)
    
    # Check if user is trying to rent their own item
    if request.user == item.owner:
        messages.error(request, "You cannot rent your own item.")
        return redirect('item_detail', pk=pk)
    
    # Check if item is available
    if not item.is_available:
        messages.error(request, "This item is not available for rent.")
        return redirect('item_detail', pk=pk)
    
    # Check if user already has a pending or active rental for this item
    existing_rental = Rental.objects.filter(
        item=item,
        borrower=request.user,
        status__in=['pending', 'approved', 'borrowed']
    ).first()
    
    if existing_rental:
        messages.warning(request, "You already have an active or pending rental request for this item.")
        return redirect('item_detail', pk=pk)

    if request.method == 'POST':
        form = RentalRequestForm(request.POST)
        if form.is_valid():
            rental = form.save(commit=False)
            rental.item = item
            rental.borrower = request.user
            rental.lender = item.owner
            rental.status = 'pending'
            rental.save()

            # Create notification for the item owner
            Notification.objects.create(
                user=item.owner,
                rental=rental,
                type='rental_request',
                message=f"{request.user.username} wants to rent your {item.name}."
            )

            messages.success(request, "Your rental request has been sent to the owner.")
            return redirect('item_detail', pk=pk)
    else:
        form = RentalRequestForm()
    
    context = {
        'form': form,
        'item': item
    }
    return render(request, 'rental_request.html', context)

@login_required(login_url='login')
def profile_view(request):
    """View user profile with their statistics"""
    user = request.user
    
    # Get user's items
    user_items = Item.objects.filter(owner=user).order_by('-date_posted')
    
    # Get rental statistics
    items_borrowed = Rental.objects.filter(borrower=user).count()
    items_lent = Rental.objects.filter(lender=user).count()
    active_borrowing = Rental.objects.filter(borrower=user, status='borrowed').count()
    active_lending = Rental.objects.filter(lender=user, status='borrowed').count()
    
    context = {
        'profile_user': user,
        'user_items': user_items,
        'items_borrowed': items_borrowed,
        'items_lent': items_lent,
        'active_borrowing': active_borrowing,
        'active_lending': active_lending,
    }
    return render(request, 'profile.html', context)

@login_required(login_url='login')
def password_change_view(request):
    """Change user password"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important to keep user logged in
            messages.success(request, 'Your password was successfully updated!')
            return redirect('profile')
    else:
        form = PasswordChangeForm(request.user)
    
    context = {'form': form}
    return render(request, 'password_change.html', context)

@login_required(login_url='login')
def edit_profile_view(request):
    """Edit user profile information"""
    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '✨ Your profile was successfully updated!')
            return redirect('profile')
    else:
        form = UserProfileEditForm(instance=request.user)
    
    context = {'form': form}
    return render(request, 'edit_profile.html', context)

@login_required(login_url='login')
def my_rented_items_view(request):
    """View items the user is currently borrowing or has borrowed"""
    # Get all rentals where user is the borrower
    my_rentals = Rental.objects.filter(borrower=request.user).select_related('item', 'lender').order_by('-request_date')
    
    # Separate by status
    active_rentals = my_rentals.filter(status='borrowed')
    pending_rentals = my_rentals.filter(status='pending')
    approved_rentals = my_rentals.filter(status='approved')
    completed_rentals = my_rentals.filter(status='returned')
    rejected_rentals = my_rentals.filter(status='rejected')
    
    context = {
        'active_rentals': active_rentals,
        'pending_rentals': pending_rentals,
        'approved_rentals': approved_rentals,
        'completed_rentals': completed_rentals,
        'rejected_rentals': rejected_rentals,
    }
    return render(request, 'my_rented_items.html', context)

@login_required(login_url='login')
def my_lended_items_view(request):
    """View items the user has lent out to others"""
    # Get all rentals where user is the lender
    my_lendings = Rental.objects.filter(lender=request.user).select_related('item', 'borrower').order_by('-request_date')
    
    # Separate by status
    pending_requests = my_lendings.filter(status='pending')
    active_lendings = my_lendings.filter(status='borrowed')
    approved_lendings = my_lendings.filter(status='approved')
    completed_lendings = my_lendings.filter(status='returned')
    rejected_lendings = my_lendings.filter(status='rejected')
    
    context = {
        'pending_requests': pending_requests,
        'active_lendings': active_lendings,
        'approved_lendings': approved_lendings,
        'completed_lendings': completed_lendings,
        'rejected_lendings': rejected_lendings,
    }
    return render(request, 'my_lended_items.html', context)

@staff_member_required
def admin_reports_dashboard(request):
    """Admin dashboard showing report statistics"""
    total_items = Item.objects.count()
    total_users = CustomUser.objects.count()
    total_rentals = Rental.objects.count()
    active_rentals = Rental.objects.filter(status='borrowed').count()
    pending_requests = Rental.objects.filter(status='pending').count()
    
    # Most borrowed items
    most_borrowed = Item.objects.annotate(
        rental_count=Count('rentals')
    ).order_by('-rental_count')[:10]
    
    # Recent activity
    recent_rentals = Rental.objects.order_by('-request_date')[:10]
    
    # User activity stats
    active_borrowers = CustomUser.objects.annotate(
        borrow_count=Count('borrowed_items')
    ).filter(borrow_count__gt=0).order_by('-borrow_count')[:10]
    
    active_lenders = CustomUser.objects.annotate(
        lend_count=Count('lent_items')
    ).filter(lend_count__gt=0).order_by('-lend_count')[:10]
    
    context = {
        'total_items': total_items,
        'total_users': total_users,
        'total_rentals': total_rentals,
        'active_rentals': active_rentals,
        'pending_requests': pending_requests,
        'most_borrowed': most_borrowed,
        'recent_rentals': recent_rentals,
        'active_borrowers': active_borrowers,
        'active_lenders': active_lenders,
    }
    return render(request, 'admin_reports.html', context)

@staff_member_required
def export_items_report_pdf(request):
    """Generate PDF report of all items"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="items_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("BroRent - Items Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary stats
    total_items = Item.objects.count()
    available_items = Item.objects.filter(is_available=True).count()
    summary_text = f"Total Items: {total_items} | Available: {available_items} | Unavailable: {total_items - available_items}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    elements.append(Spacer(1, 0.3*inch))
    
    # Items table
    items = Item.objects.all().order_by('-date_posted')
    data = [['ID', 'Name', 'Owner', 'Category', 'Price', 'Status']]
    
    for item in items:
        status = 'Available' if item.is_available else 'Unavailable'
        price_str = f"${item.price}/day" if item.per_day else f"${item.price}"
        data.append([
            str(item.pk),
            item.name[:30],
            item.owner.username,
            item.get_category_display(),
            price_str,
            status
        ])
    
    table = Table(data, colWidths=[0.5*inch, 2*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

@staff_member_required
def export_items_report_excel(request):
    """Generate Excel report of all items"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="items_report.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Items Report'
    
    # Header style
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Headers
    headers = ['ID', 'Name', 'Description', 'Owner', 'Category', 'Price', 'Per Day', 'Status', 'Date Posted']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    items = Item.objects.all().order_by('-date_posted')
    for row_num, item in enumerate(items, 2):
        worksheet.cell(row=row_num, column=1, value=item.pk)
        worksheet.cell(row=row_num, column=2, value=item.name)
        worksheet.cell(row=row_num, column=3, value=item.description)
        worksheet.cell(row=row_num, column=4, value=item.owner.username)
        worksheet.cell(row=row_num, column=5, value=item.get_category_display())
        worksheet.cell(row=row_num, column=6, value=float(item.price))
        worksheet.cell(row=row_num, column=7, value='Yes' if item.per_day else 'No')
        worksheet.cell(row=row_num, column=8, value='Available' if item.is_available else 'Unavailable')
        worksheet.cell(row=row_num, column=9, value=item.date_posted.strftime('%Y-%m-%d %H:%M'))
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    workbook.save(response)
    return response

@staff_member_required
def export_rentals_report_pdf(request):
    """Generate PDF report of most borrowed items"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rentals_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("BroRent - Rental Activity Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary stats
    total_rentals = Rental.objects.count()
    active_rentals = Rental.objects.filter(status='borrowed').count()
    completed_rentals = Rental.objects.filter(status='returned').count()
    summary_text = f"Total Rentals: {total_rentals} | Active: {active_rentals} | Completed: {completed_rentals}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    elements.append(Spacer(1, 0.3*inch))
    
    # Most borrowed items
    subtitle = Paragraph("Most Borrowed Items", styles['Heading2'])
    elements.append(subtitle)
    elements.append(Spacer(1, 0.1*inch))
    
    most_borrowed = Item.objects.annotate(
        rental_count=Count('rentals')
    ).order_by('-rental_count')[:15]
    
    data = [['Rank', 'Item Name', 'Owner', 'Times Borrowed', 'Category']]
    for idx, item in enumerate(most_borrowed, 1):
        data.append([
            str(idx),
            item.name[:30],
            item.owner.username,
            str(item.rental_count),
            item.get_category_display()
        ])
    
    table = Table(data, colWidths=[0.6*inch, 2.5*inch, 1.2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

@staff_member_required
def export_rentals_report_excel(request):
    """Generate Excel report of rental activity"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="rentals_report.xlsx"'
    
    workbook = Workbook()
    
    # Sheet 1: All Rentals
    worksheet1 = workbook.active
    worksheet1.title = 'All Rentals'
    
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    headers1 = ['ID', 'Item', 'Borrower', 'Lender', 'Status', 'Start Date', 'End Date', 'Total Price', 'Request Date']
    for col_num, header in enumerate(headers1, 1):
        cell = worksheet1.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    rentals = Rental.objects.all().order_by('-request_date')
    for row_num, rental in enumerate(rentals, 2):
        worksheet1.cell(row=row_num, column=1, value=rental.pk)
        worksheet1.cell(row=row_num, column=2, value=rental.item.name)
        worksheet1.cell(row=row_num, column=3, value=rental.borrower.username)
        worksheet1.cell(row=row_num, column=4, value=rental.lender.username)
        worksheet1.cell(row=row_num, column=5, value=rental.get_status_display())
        worksheet1.cell(row=row_num, column=6, value=rental.start_date.strftime('%Y-%m-%d'))
        worksheet1.cell(row=row_num, column=7, value=rental.end_date.strftime('%Y-%m-%d'))
        worksheet1.cell(row=row_num, column=8, value=float(rental.total_price) if rental.total_price else 0)
        worksheet1.cell(row=row_num, column=9, value=rental.request_date.strftime('%Y-%m-%d %H:%M'))
    
    for col in range(1, len(headers1) + 1):
        worksheet1.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 2: Most Borrowed Items
    worksheet2 = workbook.create_sheet(title='Most Borrowed')
    
    headers2 = ['Rank', 'Item Name', 'Owner', 'Category', 'Times Borrowed', 'Total Revenue']
    for col_num, header in enumerate(headers2, 1):
        cell = worksheet2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    most_borrowed = Item.objects.annotate(
        rental_count=Count('rentals')
    ).order_by('-rental_count')[:20]
    
    for row_num, item in enumerate(most_borrowed, 2):
        total_revenue = sum(
            float(r.total_price) if r.total_price else 0 
            for r in item.rentals.filter(status__in=['borrowed', 'returned'])
        )
        worksheet2.cell(row=row_num, column=1, value=row_num - 1)
        worksheet2.cell(row=row_num, column=2, value=item.name)
        worksheet2.cell(row=row_num, column=3, value=item.owner.username)
        worksheet2.cell(row=row_num, column=4, value=item.get_category_display())
        worksheet2.cell(row=row_num, column=5, value=item.rental_count)
        worksheet2.cell(row=row_num, column=6, value=total_revenue)
    
    for col in range(1, len(headers2) + 1):
        worksheet2.column_dimensions[get_column_letter(col)].width = 18
    
    workbook.save(response)
    return response

@staff_member_required
def export_users_report_pdf(request):
    """Generate PDF report of user activity"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="users_activity_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("BroRent - User Activity Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary
    total_users = CustomUser.objects.count()
    active_borrowers = CustomUser.objects.annotate(
        borrow_count=Count('borrowed_items')
    ).filter(borrow_count__gt=0).count()
    
    active_lenders = CustomUser.objects.annotate(
        lend_count=Count('lent_items')
    ).filter(lend_count__gt=0).count()
    
    summary_text = f"Total Users: {total_users} | Active Borrowers: {active_borrowers} | Active Lenders: {active_lenders}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    elements.append(Spacer(1, 0.3*inch))
    
    # Top borrowers
    subtitle1 = Paragraph("Top Borrowers", styles['Heading2'])
    elements.append(subtitle1)
    elements.append(Spacer(1, 0.1*inch))
    
    top_borrowers = CustomUser.objects.annotate(
        borrow_count=Count('borrowed_items')
    ).filter(borrow_count__gt=0).order_by('-borrow_count')[:10]
    
    data1 = [['Rank', 'Username', 'Hostel', 'Items Borrowed', 'Phone']]
    for idx, user in enumerate(top_borrowers, 1):
        data1.append([
            str(idx),
            user.username,
            user.hostel_name,
            str(user.borrow_count),
            user.phone_number
        ])
    
    table1 = Table(data1, colWidths=[0.6*inch, 1.5*inch, 1.5*inch, 1.2*inch, 1.5*inch])
    table1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table1)
    elements.append(Spacer(1, 0.3*inch))
    
    # Top lenders
    subtitle2 = Paragraph("Top Lenders", styles['Heading2'])
    elements.append(subtitle2)
    elements.append(Spacer(1, 0.1*inch))
    
    top_lenders = CustomUser.objects.annotate(
        lend_count=Count('lent_items'),
        items_listed=Count('item')
    ).filter(lend_count__gt=0).order_by('-lend_count')[:10]
    
    data2 = [['Rank', 'Username', 'Hostel', 'Times Lent', 'Items Listed']]
    for idx, user in enumerate(top_lenders, 1):
        data2.append([
            str(idx),
            user.username,
            user.hostel_name,
            str(user.lend_count),
            str(user.items_listed)
        ])
    
    table2 = Table(data2, colWidths=[0.6*inch, 1.5*inch, 1.5*inch, 1.2*inch, 1.2*inch])
    table2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table2)
    
    doc.build(elements)
    return response

@staff_member_required
def export_users_report_excel(request):
    """Generate Excel report of user activity"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="users_activity_report.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'User Activity'
    
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    headers = ['ID', 'Username', 'Hostel', 'Room', 'Phone', 'Items Listed', 'Times Borrowed', 'Times Lent', 'Date Joined']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    users = CustomUser.objects.annotate(
        items_count=Count('item'),
        borrow_count=Count('borrowed_items'),
        lend_count=Count('lent_items')
    ).order_by('-borrow_count', '-lend_count')
    
    for row_num, user in enumerate(users, 2):
        worksheet.cell(row=row_num, column=1, value=user.pk)
        worksheet.cell(row=row_num, column=2, value=user.username)
        worksheet.cell(row=row_num, column=3, value=user.hostel_name)
        worksheet.cell(row=row_num, column=4, value=user.room_number)
        worksheet.cell(row=row_num, column=5, value=user.phone_number)
        worksheet.cell(row=row_num, column=6, value=user.items_count)
        worksheet.cell(row=row_num, column=7, value=user.borrow_count)
        worksheet.cell(row=row_num, column=8, value=user.lend_count)
        worksheet.cell(row=row_num, column=9, value=user.date_joined.strftime('%Y-%m-%d'))
    
    for col in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    workbook.save(response)
    return response

@login_required(login_url='login')
def notifications_list_view(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    context = {
        'notifications': notifications
    }
    return render(request, 'notifications.html', context)

@login_required(login_url='login')
def notification_mark_read_view(request, pk):
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    if not notification.read:
        notification.read = True
        notification.save()
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('notifications')

@login_required(login_url='login')
def rental_request_accept_view(request, pk):
    """Accept a rental request"""
    rental = get_object_or_404(Rental, pk=pk)
    
    # Check if the current user is the lender (item owner)
    if request.user != rental.lender:
        messages.error(request, "You don't have permission to accept this request.")
        return redirect('notifications')
    
    # Check if the rental is still pending
    if rental.status != 'pending':
        messages.error(request, "This rental request has already been processed.")
        return redirect('notifications')
    
    # Accept the rental request
    rental.status = 'approved'
    rental.approved_date = timezone.now()
    rental.save()
    
    # Create notification for the borrower
    Notification.objects.create(
        user=rental.borrower,
        rental=rental,
        type='request_approved',
        message=f"Your rental request for {rental.item.name} has been approved!"
    )
    
    messages.success(request, f"Rental request from {rental.borrower.username} has been approved.")
    return redirect('notifications')

@login_required(login_url='login')
def rental_request_reject_view(request, pk):
    """Reject a rental request"""
    rental = get_object_or_404(Rental, pk=pk)
    
    # Check if the current user is the lender (item owner)
    if request.user != rental.lender:
        messages.error(request, "You don't have permission to reject this request.")
        return redirect('notifications')
    
    # Check if the rental is still pending
    if rental.status != 'pending':
        messages.error(request, "This rental request has already been processed.")
        return redirect('notifications')
    
    # Reject the rental request
    rental.status = 'rejected'
    rental.save()
    
    # Create notification for the borrower
    Notification.objects.create(
        user=rental.borrower,
        rental=rental,
        type='request_rejected',
        message=f"Your rental request for {rental.item.name} has been rejected."
    )
    
    messages.success(request, f"Rental request from {rental.borrower.username} has been rejected.")
    return redirect('notifications')